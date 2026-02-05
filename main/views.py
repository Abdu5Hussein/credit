from django.shortcuts import render
from django.http import HttpResponse
import io, re, os, zipfile
import pandas as pd
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from typing import Dict, Optional, Tuple, List
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

def read_dlt_text(uploaded_file) -> str:
        """
        Robustly decode DLT text:
        - handles UTF-8/UTF-8-SIG
        - handles UTF-16 with BOM (common in some exporters)
        """
        b = uploaded_file.read()

        # Fast BOM checks
        if b.startswith(b"\xff\xfe") or b.startswith(b"\xfe\xff"):
            return b.decode("utf-16")  # auto handles endianness from BOM

        # Try UTF-8-sig first
        try:
            s = b.decode("utf-8-sig")
            # If it contains many NULs, it's probably UTF-16 decoded wrongly somewhere upstream
            if "\x00" in s:
                return b.decode("utf-16")
            return s
        except UnicodeDecodeError:
            pass

        # Fallbacks
        for enc in ("utf-16", "cp1256", "latin-1"):
            try:
                return b.decode(enc)
            except UnicodeDecodeError:
                continue

        # Last resort
        return b.decode("utf-8", errors="replace")


class IffDltToExcelAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        dlt_file = request.FILES.get("dlt_file")
        if not dlt_file:
            return Response({"detail": "Upload dlt_file"}, status=status.HTTP_400_BAD_REQUEST)

        # Read text
        raw = dlt_file.read().decode("utf-8", errors="replace")
        lines = raw.splitlines()

        # Choose max fields (increase if your DLT is wider)
        max_fields = int(request.data.get("max_fields", 60))

        def split_fields(line: str):
            parts = line.split("|")
            parts += [""] * (max_fields - len(parts))
            return parts[:max_fields]

        rows = []
        for i, line in enumerate(lines, start=1):
            parts = split_fields(line)
            seg = parts[0] if parts else ""
            recid = None
            for p in parts:
                if isinstance(p, str) and re.fullmatch(r"LD\d+", p):
                    recid = p
                    break

            row = {"LINE_NO": i, "SEGMENT": seg, "RECID": recid, "RAW_LINE": line}
            for idx, val in enumerate(parts, start=1):
                row[f"F{idx:02d}"] = val
            rows.append(row)

        df = pd.DataFrame(rows)

        # Write excel in-memory
        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="DLT", index=False)

        # Light formatting
        out.seek(0)
        wb = load_workbook(out)
        ws = wb["DLT"]
        ws.freeze_panes = "A2"

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for c in ws[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[:200]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 55)

        final = io.BytesIO()
        wb.save(final)
        final.seek(0)

        base = os.path.splitext(dlt_file.name)[0]
        safe = re.sub(r"[^A-Za-z0-9._-]+", "_", base)
        filename = f"IFF_DLT-{safe}.xlsx"

        resp = HttpResponse(
            final.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

class ExcelToIffDltAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        xlsx_file = request.FILES.get("excel_file")
        sheet = request.data.get("sheet", "DLT")  # default sheet name used above
        if not xlsx_file:
            return Response({"detail": "Upload excel_file"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(xlsx_file, sheet_name=sheet, dtype=str).fillna("")
        except Exception as e:
            return Response({"detail": f"Failed to read Excel: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        if "SEGMENT" not in df.columns:
            return Response({"detail": "Excel must contain a SEGMENT column"}, status=status.HTTP_400_BAD_REQUEST)

        # Collect Fxx columns
        fcols = [c for c in df.columns if re.fullmatch(r"F\d{2}", str(c))]
        fcols = sorted(fcols)  # F01..F99

        # Build DLT lines
        out_lines = []
        for _, row in df.iterrows():
            seg = str(row.get("SEGMENT", "")).strip()
            if not seg:
                continue

            fields = []
            for c in fcols:
                fields.append(str(row.get(c, "")).strip())

            # trim trailing empty fields
            while fields and fields[-1] == "":
                fields.pop()

            out_lines.append("|".join(fields))

        dlt_text = "\n".join(out_lines) + ("\n" if out_lines else "")

        base = os.path.splitext(xlsx_file.name)[0]
        safe = re.sub(r"[^A-Za-z0-9._-]+", "_", base)
        filename = f"IFF_DLT-{safe}.dlt"

        resp = HttpResponse(dlt_text, content_type="text/plain; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


class DltErrorToExcelAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        dlt = request.FILES.get("dlt_file")
        err = request.FILES.get("error_log")

        if not dlt or not err:
            return Response(
                {"detail": "Upload both: dlt_file and error_log"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 1) Read error log (pipe-delimited)
        try:
            df_err = pd.read_csv(err, sep="|", encoding="utf-8-sig")
        except Exception as e:
            return Response({"detail": f"Bad error_log CSV: {e}"}, status=400)

        if "LINE_NUMBER" in df_err.columns:
            df_err["LINE_NUMBER"] = pd.to_numeric(df_err["LINE_NUMBER"], errors="coerce").astype("Int64")
        else:
            df_err["LINE_NUMBER"] = pd.Series([pd.NA] * len(df_err), dtype="Int64")

        # 2) Read DLT file
        raw = dlt.read().decode("utf-8", errors="replace")
        lines = raw.splitlines()

        max_fields = 60

        def split_fields(line: str):
            parts = line.split("|")
            parts += [""] * (max_fields - len(parts))
            return parts[:max_fields]

        dlt_rows = []
        for i, line in enumerate(lines, start=1):
            parts = split_fields(line)
            seg = parts[0] if parts else ""
            recid = None
            for p in parts:
                if isinstance(p, str) and p.startswith("LD") and re.fullmatch(r"LD\d+", p):
                    recid = p
                    break

            row = {"DLT_LINE": i, "SEGMENT": seg, "RECID": recid, "RAW_LINE": line}
            for idx, val in enumerate(parts, start=1):
                row[f"F{idx:02d}"] = val
            dlt_rows.append(row)

        df_dlt = pd.DataFrame(dlt_rows)

        # 3) Fill missing LINE_NUMBER using RECORD_REFERENCE_VALUE (LDxxxx) if present
        recid_first = (
            df_dlt.dropna(subset=["RECID"])
            .groupby("RECID")["DLT_LINE"]
            .min()
        )

        def fill_line(row):
            if pd.isna(row.get("LINE_NUMBER")):
                rec = row.get("RECORD_REFERENCE_VALUE")
                if isinstance(rec, str) and rec in recid_first.index:
                    return int(recid_first.loc[rec])
            return row.get("LINE_NUMBER")

        df_err["LINE_NUMBER_FILLED"] = df_err.apply(fill_line, axis=1).astype("Int64")

        # 4) Join
        df_join = df_err.merge(df_dlt, how="left", left_on="LINE_NUMBER_FILLED", right_on="DLT_LINE")

        df_join["ERROR_SHORT"] = (
            df_join.get("ERROR_ID", "").astype(str)
            + " | " + df_join.get("ERROR_SEVERITY", "").astype(str)
            + " | " + df_join.get("ERROR_FIELD_NAME", "").fillna("").astype(str)
            + " | " + df_join.get("ERROR_MESSAGE", "").fillna("").astype(str)
        )

        df_join["DLT_KEY"] = (
            df_join.get("SEGMENT", "").fillna("").astype(str)
            + " | " + df_join.get("RECID", "").fillna("").astype(str)
        )

        # 5) Summary
        summary_sev = (
            df_err.groupby("ERROR_SEVERITY").size().reset_index(name="Count")
            if "ERROR_SEVERITY" in df_err.columns else pd.DataFrame()
        )

        summary_id = (
            df_err.groupby(["ERROR_ID", "ERROR_SEVERITY"]).size().reset_index(name="Count")
            if ("ERROR_ID" in df_err.columns and "ERROR_SEVERITY" in df_err.columns)
            else pd.DataFrame()
        )

        df_dlt_view = df_dlt.copy()
        df_dlt_view["HAS_RECID"] = df_dlt_view["RECID"].notna()
        df_dlt_view = df_dlt_view.sort_values(["HAS_RECID", "DLT_LINE"], ascending=[False, True])

        # 6) Write Excel to memory
        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df_join.to_excel(writer, sheet_name="Joined (Per Error)", index=False)
            df_err.to_excel(writer, sheet_name="Errors (Raw)", index=False)
            df_dlt_view.to_excel(writer, sheet_name="DLT (Parsed)", index=False)
            if not summary_sev.empty or not summary_id.empty:
                summary_sev.to_excel(writer, sheet_name="Summary", index=False, startrow=0)
                summary_id.to_excel(writer, sheet_name="Summary", index=False, startrow=len(summary_sev) + 3)

        # 7) Formatting
        out.seek(0)
        wb = load_workbook(out)

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)

        for name in wb.sheetnames:
            ws = wb[name]
            ws.freeze_panes = "A2"
            for c in ws[1]:
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col[:200]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 55)

        if "Joined (Per Error)" in wb.sheetnames:
            ws = wb["Joined (Per Error)"]
            header = [c.value for c in ws[1]]
            if "ERROR_SEVERITY" in header:
                sev_col = header.index("ERROR_SEVERITY") + 1

                # Criticality shades
                structure_red = PatternFill("solid", fgColor="991B1B")  # very dark red
                segment_red   = PatternFill("solid", fgColor="DC2626")  # strong red
                entity_red    = PatternFill("solid", fgColor="FCA5A5")  # soft red
                field_yellow  = PatternFill("solid", fgColor="FDE68A")  # yellow

                # Optional (if your file has these severities)
                warning_or_other = PatternFill("solid", fgColor="E5E7EB")  # light gray

                for r in range(2, ws.max_row + 1):
                    val = (ws.cell(r, sev_col).value or "").strip()

                    if val == "Structure Rejection":
                        ws.cell(r, sev_col).fill = structure_red
                    elif val == "Segment Rejection":
                        ws.cell(r, sev_col).fill = segment_red
                    elif val == "Entity Rejection":
                        ws.cell(r, sev_col).fill = entity_red
                    elif val == "Field Rejection":
                        ws.cell(r, sev_col).fill = field_yellow
                    else:
                        # optional: color anything else lightly
                        # ws.cell(r, sev_col).fill = warning_or_other
                        pass


        final = io.BytesIO()
        wb.save(final)
        final.seek(0)

        resp = HttpResponse(
            final.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        original = dlt.name  # e.g. "2C4L20020-COM-17-Dec-2025-235900.dlt"
        base = os.path.splitext(original)[0]  # remove .dlt
        safe = re.sub(r'[^A-Za-z0-9._-]+', '_', base)  # replace spaces/special chars
        resp["Content-Disposition"] = f'attachment; filename="DLT_Error_Analysis-{safe}.xlsx"'
        return resp

class ManuallyCloseCommercialDLT(APIView):
    """
    Upload:
      - dlt_file فقط

    Safety rules:
      - DLT must contain exactly ONE CMCF and ONE CMCS (one pair only)
      - Both must share the same RECID (LDxxxx)
      - Overwrite ONLY the static closure fields (12 fields from your template)
      - Fail if any of the 12 codes is missing in the CMCF line (strict & safe)
    """
    parser_classes = (MultiPartParser, FormParser)

    MAX_FIELDS_PADDING = 400  # padding for safer scanning

    # -----------------------------
    # STATIC CLOSURE VALUES (12)
    # -----------------------------
    # Values derived from your uploaded template:
    # - CMCF22 => 003
    # - CMCF23 => 001
    # - CMCF19/20/21/24/25/38/40 => 0
    # - CMCF39 => EMPTY
    # - CMCF37 => pick one valid C15 code (fixed here: 001)
    #
    # NOTE: CMCF35 must be a valid date. We'll set it to today's date (server time),
    # formatted as DD-MMM-YYYY with English month abbreviations.
    CLOSURE_FIELDS_STATIC: Dict[str, str] = {
        "CMCF22": "003",
        "CMCF19": "0",
        "CMCF20": "0",
        "CMCF21": "0",
        "CMCF23": "001",
        "CMCF24": "0",
        "CMCF25": "0",
        "CMCF38": "0",
        "CMCF39": "",
        "CMCF40": "0",
        "CMCF37": "001",
        # CMCF35 is handled dynamically (today) in _closure_overwrite_map()
    }

    def post(self, request):
        dlt = request.FILES.get("dlt_file")
        if not dlt:
            return Response({"detail": "Upload: dlt_file"}, status=status.HTTP_400_BAD_REQUEST)

        raw = read_dlt_text(dlt)
        lines = raw.splitlines()

        # 1) Safety: exactly one CMCS + one CMCF
        cmcs_line_no, cmcs_recid_or_err = self._find_single_segment(lines, "CMCS")
        if cmcs_line_no is None:
            return Response(cmcs_recid_or_err, status=status.HTTP_400_BAD_REQUEST)

        cmcf_line_no, cmcf_recid_or_err = self._find_single_segment(lines, "CMCF")
        if cmcf_line_no is None:
            return Response(cmcf_recid_or_err, status=status.HTTP_400_BAD_REQUEST)

        cmcs_recid = cmcs_recid_or_err
        cmcf_recid = cmcf_recid_or_err

        if not cmcs_recid or not cmcf_recid:
            return Response(
                {"detail": "Could not detect RECID (LDxxxx) in CMCS/CMCF line."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if cmcs_recid != cmcf_recid:
            return Response(
                {
                    "detail": "Safety validation failed: CMCS and CMCF do not belong to the same RECID.",
                    "cmcs_recid": cmcs_recid,
                    "cmcf_recid": cmcf_recid,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 2) Apply closure overwrites ONLY to the CMCF line (fields are CMCFxx)
        overwrite_map = self._closure_overwrite_map()  # contains CMCF19..CMCF40 etc

        new_lines = lines[:]

        for i, line in enumerate(new_lines):
            if line.startswith("HDHD|") or line.strip() == "HDHD":
                new_lines[i] = self._set_hdhd_dates_to_today(line)
                break  # only one header expected

        new_cmcf_line, overwritten_codes, missing_codes = self._overwrite_by_position_strict(
            new_lines[cmcf_line_no - 1],
            overwrite_map,
        )

        if missing_codes:
            return Response(
                {
                    "detail": "Safety check failed: CMCF record doesn't contain enough positional fields for closure.",
                    "missing_codes": missing_codes,
                    "hint": "This DLT line has fewer columns than expected. Confirm the CMCF schema/field count.",
                },
                status=400,
            )

        new_lines[cmcf_line_no - 1] = new_cmcf_line

        # 3) Strict safety: ensure all 12+1 fields exist in CMCF record and were overwritten
        missing = sorted(set(overwrite_map.keys()) - set(overwritten_codes))
        if missing:
            return Response(
                {
                    "detail": "Safety check failed: some closure field codes were not found in the single CMCF record line.",
                    "missing_codes": missing,
                    "hint": "This DLT record structure differs from expected schema OR some fields are not present in the record.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        out_text = "\n".join(new_lines) + ("\n" if raw.endswith("\n") else "")

        original_name = dlt.name
        base = os.path.splitext(original_name)[0]
        safe = re.sub(r"[^A-Za-z0-9._-]+", "_", base)

        # Find HDHD line number for reporting (after modifications it should still exist)
        hdhd_line_no = self._find_first_segment_line_no(lines, "HDHD")

        excel_bytes = self._build_change_excel(
            original_lines=lines,     # original
            new_lines=new_lines,      # modified
            hdhd_line_no=hdhd_line_no,
            cmcs_line_no=cmcs_line_no,
            cmcf_line_no=cmcf_line_no,
        )

        # Build ZIP: closed dlt + excel report
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
            z.writestr(f"CLOSED_{safe}.dlt", out_text.encode("utf-8"))
            z.writestr(f"DLT_CHANGES_{safe}.xlsx", excel_bytes)

        zip_buf.seek(0)

        resp = HttpResponse(zip_buf.getvalue(), content_type="application/zip")
        resp["Content-Disposition"] = f'attachment; filename="CLOSED_{safe}_WITH_REPORT.zip"'
        return resp

    # ---------------- helpers ----------------

    def _closure_overwrite_map(self) -> Dict[str, str]:
        """
        Build final overwrite map for closure:
        includes the 12 static fields + CMCF35 = today date (DD-MMM-YYYY).
        """
        m = dict(self.CLOSURE_FIELDS_STATIC)
        m["CMCF35"] = self._today_dd_mmm_yyyy()
        return m

    def _set_hdhd_dates_to_today(self, line: str) -> str:
        """
        HDHD format (positional):
        HDHD|<ref>|<date1>|<date2>|<time>|<seq>
        We overwrite date1 (parts[2]) and date2 (parts[3]) with today's date.
        """
        parts = line.split("|")
        if not parts or parts[0].strip() != "HDHD":
            return line

        today = self._today_dd_mmm_yyyy()

        # Ensure indexes exist
        if len(parts) > 2:
            parts[2] = today
        if len(parts) > 3:
            parts[3] = today

        return "|".join(parts)

    def _today_dd_mmm_yyyy(self) -> str:
        """
        Ensure English MMM abbreviations regardless of server locale.
        """
        now = datetime.now()
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        return f"{now.day:02d}-{months[now.month - 1]}-{now.year}"

    def _find_single_segment(self, lines: List[str], segment: str) -> Tuple[Optional[int], object]:
        """
        Returns (line_no, recid) OR (None, error_payload_dict)
        """
        hits = []
        for i, line in enumerate(lines, start=1):
            parts = line.split("|")
            seg = parts[0].strip() if parts else ""
            if seg == segment:
                recid = self._extract_recid(parts)
                hits.append((i, recid))

        if len(hits) != 1:
            return None, {
                "detail": f"Safety validation failed: DLT must contain exactly ONE {segment} record.",
                f"{segment.lower()}_count": len(hits),
                f"{segment.lower()}_lines": [h[0] for h in hits],
            }

        return hits[0][0], hits[0][1]

    def _extract_recid(self, parts: List[str]) -> Optional[str]:
        for p in parts:
            p = (p or "").strip()
            if p.startswith("LD") and re.fullmatch(r"LD\d+", p):
                return p
        return None

    def _build_change_excel(
        self,
        original_lines: List[str],
        new_lines: List[str],
        hdhd_line_no: Optional[int],
        cmcs_line_no: int,
        cmcf_line_no: int,
    ) -> bytes:
        """
        Creates an Excel file comparing original vs new.
        - Summary sheet: which lines changed
        - HDHD diff: field-by-field (positional)
        - CMCF diff: field-by-field (positional + highlights only changed fields)
        - CMCS diff (optional): field-by-field if you ever overwrite CMCS later
        """
        wb = Workbook()

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        changed_fill = PatternFill("solid", fgColor="FDE68A")  # highlight changed
        mono = Font(name="Consolas")

        def style_header(ws):
            ws.freeze_panes = "A2"
            for c in ws[1]:
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def autosize(ws, max_rows=200):
            for col in ws.columns:
                max_len = 0
                letter = get_column_letter(col[0].column)
                for cell in col[:max_rows]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[letter].width = min(max(12, max_len + 2), 60)

        def add_line_sheet(title: str, line_no: int):
            ws = wb.create_sheet(title)
            ws.append(["Line No", "Original Line", "New Line", "Changed?"])
            style_header(ws)

            o = original_lines[line_no - 1] if 0 < line_no <= len(original_lines) else ""
            n = new_lines[line_no - 1] if 0 < line_no <= len(new_lines) else ""
            changed = (o != n)

            ws.append([line_no, o, n, "YES" if changed else "NO"])
            ws["B2"].font = mono
            ws["C2"].font = mono
            if changed:
                ws["D2"].fill = changed_fill
            autosize(ws)

        def add_positional_diff_sheet(title: str, line_no: int, segment: str):
            ws = wb.create_sheet(title)
            ws.append(["FieldNo", "Index", "Original", "New", "Changed?"])
            style_header(ws)

            o = original_lines[line_no - 1].split("|") if 0 < line_no <= len(original_lines) else []
            n = new_lines[line_no - 1].split("|") if 0 < line_no <= len(new_lines) else []

            # Only proceed if the segment matches
            if (o and o[0].strip() != segment) or (n and n[0].strip() != segment):
                ws.append(["-", "-", f"Segment mismatch. Expected {segment}.", "", ""])
                autosize(ws)
                return

            max_len = max(len(o), len(n))
            for idx in range(max_len):
                field_no = idx + 1  # field numbering includes segment token
                ov = o[idx] if idx < len(o) else ""
                nv = n[idx] if idx < len(n) else ""
                changed = (ov != nv)

                ws.append([field_no, idx, ov, nv, "YES" if changed else "NO"])

                # monospace for values
                ws.cell(row=ws.max_row, column=3).font = mono
                ws.cell(row=ws.max_row, column=4).font = mono

                if changed:
                    ws.cell(row=ws.max_row, column=3).fill = changed_fill
                    ws.cell(row=ws.max_row, column=4).fill = changed_fill
                    ws.cell(row=ws.max_row, column=5).fill = changed_fill

            autosize(ws)

        # Remove default sheet
        wb.remove(wb.active)

        # Summary
        ws0 = wb.create_sheet("Summary")
        ws0.append(["Line Type", "Line No", "Changed?", "Notes"])
        style_header(ws0)

        def add_summary_row(label, ln, note=""):
            o = original_lines[ln - 1] if 0 < ln <= len(original_lines) else ""
            n = new_lines[ln - 1] if 0 < ln <= len(new_lines) else ""
            changed = "YES" if o != n else "NO"
            ws0.append([label, ln, changed, note])
            if changed == "YES":
                ws0.cell(row=ws0.max_row, column=3).fill = changed_fill

        if hdhd_line_no:
            add_summary_row("HDHD", hdhd_line_no, "Header dates updated to today (fields 3 & 4).")
        add_summary_row("CMCS", cmcs_line_no, "Validated only (not overwritten).")
        add_summary_row("CMCF", cmcf_line_no, "Closure fields overwritten (positional).")

        autosize(ws0)

        # Detailed sheets
        if hdhd_line_no:
            add_line_sheet("HDHD Line", hdhd_line_no)
            add_positional_diff_sheet("HDHD Diff", hdhd_line_no, "HDHD")

        add_line_sheet("CMCF Line", cmcf_line_no)
        add_positional_diff_sheet("CMCF Diff", cmcf_line_no, "CMCF")

        add_line_sheet("CMCS Line", cmcs_line_no)
        add_positional_diff_sheet("CMCS Diff", cmcs_line_no, "CMCS")

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()


    def _find_first_segment_line_no(self, lines: List[str], segment: str) -> Optional[int]:
        for i, line in enumerate(lines, start=1):
            if (line.split("|")[0].strip() if line else "") == segment:
                return i
        return None

    def _overwrite_by_position_strict(self, line: str, overwrites: Dict[str, str]) -> Tuple[str, List[str], List[str]]:
        """
        Positional overwrite:
        - Line format: CMCF|v2|v3|... (NOT TAG|VALUE pairs)
        - Field numbering includes the segment token:
            field 1 -> parts[0] == "CMCF"
            field N -> parts[N-1]

        Strict:
        - will FAIL (via missing_codes) if line doesn't have enough fields to reach a required position
        - will NOT change parts[0] (segment)
        """
        parts = line.split("|")
        seg = (parts[0] if parts else "").strip()

        overwritten = []
        missing = []

        for code, val in overwrites.items():
            code = str(code).strip()
            m = re.match(r"^(CMCF|CMCS)(\d+)$", code)
            if not m:
                continue

            expected_seg = m.group(1)
            pos = int(m.group(2))  # e.g. 22

            # Only apply matching segment
            if expected_seg != seg:
                continue

            idx = pos - 1  # because parts[0] is field 1

            # protect segment token
            if idx == 0:
                continue

            if idx >= len(parts):
                missing.append(code)
                continue

            parts[idx] = str(val)
            overwritten.append(code)

        return "|".join(parts), overwritten, missing

# Create your views here.
def home(request):
    return render(request, "main/home.html")